#!/usr/bin/env python3
"""Parse PHYSIOTHERAPY DEPT DATA new.xlsx into physio-seed-data.json."""

from __future__ import annotations

import json
import re
import uuid
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "PHYSIOTHERAPY DEPT DATA new.xlsx"
OUT = Path(__file__).resolve().parent / "physio-seed-data.json"

DAYS = {
    "monday": 1,
    "tuesday": 2,
    "wednesday": 3,
    "thursday": 4,
    "friday": 5,
    "saturday": 6,
}

FACULTY_EMAIL = {
    "RG": "riya.gupta@mygyanvihar.com",
    "PB": "prachi.baheti@mygyanvihar.com",
    "AS": "ajit.surana@mygyanvihar.com",
    "HOD": "gaurav.agarwal@mygyanvihar.com",
    # Batch-sheet faculty — mapped to HOD until dedicated accounts exist
    "AD": "gaurav.agarwal@mygyanvihar.com",
    "HM": "gaurav.agarwal@mygyanvihar.com",
    "SR": "gaurav.agarwal@mygyanvihar.com",
    "SS": "gaurav.agarwal@mygyanvihar.com",
    "VS": "gaurav.agarwal@mygyanvihar.com",
    "NF": "gaurav.agarwal@mygyanvihar.com",
    "NF 1": "gaurav.agarwal@mygyanvihar.com",
    "NF 2": "gaurav.agarwal@mygyanvihar.com",
}

SKIP_CELL = re.compile(
    r"^(LUNCH|L\s*U\s*N\s*C\s*H|PT\s*MEETING|LIBRARY|SATURDAY\s*ACTIVITY|DAY\.?\.?\.?|DAY)$",
    re.I,
)

FACULTY_TAG = re.compile(r"\(([A-Z]{1,2}(?:\s*\d)?)\)?\s*$", re.I)
TIME_RANGE = re.compile(
    r"(\d{1,2})[:.](\d{2})\s*[-–]\s*(\d{1,2})[:.](\d{2})"
)


def norm_time(h: str, m: str) -> str:
    return f"{int(h):02d}:{m}"


def parse_time_header(cell: str | None) -> tuple[str, str] | None:
    if not cell or not isinstance(cell, str):
        return None
    flat = " ".join(cell.split())
    m = TIME_RANGE.search(flat.replace("\n", " "))
    if not m:
        return None
    return norm_time(m.group(1), m.group(2)), norm_time(m.group(3), m.group(4))


def extract_faculty_key(text: str) -> str | None:
    m = FACULTY_TAG.search(text.strip())
    if not m:
        return None
    return m.group(1).upper().replace("  ", " ")


def clean_course_name(text: str) -> str:
    name = FACULTY_TAG.sub("", text).strip()
    name = re.sub(r"\s+", " ", name)
    return name


def is_skip_slot(text: str) -> bool:
    upper = text.upper().strip()
    if not upper or SKIP_CELL.match(upper):
        return True
    if upper in {"L U N C H", "L\n U\n N\n C\n H"}:
        return True
    for kw in ("PT MEETING", "LIBRARY", "SATURDAY ACTIVITY"):
        if kw in upper:
            return True
    return False


def slug_code(name: str, sem_roman: str, existing: set[str]) -> str:
    base = re.sub(r"[^A-Z0-9]", "", name.upper())[:12] or "COURSE"
    kind = "P" if any(x in name.upper() for x in ("LAB", "PRACTICAL", "PRACT")) else "T"
    prefix = {"III": "3", "V": "5", "VII": "7"}.get(sem_roman, "0")
    code = f"BPT{prefix}{base[:6]}{kind}"
    if len(code) > 12:
        code = code[:12]
    n = 1
    candidate = code
    while candidate in existing:
        n += 1
        candidate = f"{code[:10]}{n:02d}"
    existing.add(candidate)
    return candidate


def parse_batch_sheet(ws, sem_roman: str, room: str, section: str) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    time_cols: list[tuple[int, str, str]] = []
    for i, row in enumerate(rows):
        if row and isinstance(row[0], str) and row[0].strip().lower().startswith("day"):
            header_idx = i
            for col_idx, cell in enumerate(row):
                if col_idx == 0:
                    continue
                parsed = parse_time_header(str(cell) if cell else None)
                if parsed:
                    time_cols.append((col_idx, parsed[0], parsed[1]))
            break
    if header_idx is None:
        raise ValueError(f"Could not find header row in {ws.title}")

    courses: dict[str, dict] = {}
    code_set: set[str] = set()
    slots: list[dict] = []
    students: list[dict] = []

    for row in rows[header_idx + 1 :]:
        if not row or not row[0]:
            continue
        day_cell = str(row[0]).strip().lower()
        if day_cell not in DAYS:
            continue
        day = DAYS[day_cell]
        for col_idx, start, end in time_cols:
            if col_idx >= len(row):
                continue
            cell = row[col_idx]
            if not cell or not isinstance(cell, str):
                continue
            text = cell.strip()
            if is_skip_slot(text):
                continue
            name = clean_course_name(text)
            if not name:
                continue
            fkey = extract_faculty_key(text) or "HOD"
            fkey = fkey.upper()
            email_key = fkey if fkey in FACULTY_EMAIL else "HOD"
            code = slug_code(name, sem_roman, code_set)
            ctype = "LAB" if kind_is_lab(name) else "THEORY"
            credits = 2 if ctype == "LAB" else 4
            courses[code] = {
                "code": code,
                "name": name,
                "credits": credits,
                "type": ctype,
                "faculty": email_key,
            }
            slots.append(
                {
                    "course": code,
                    "day": day,
                    "start": start,
                    "end": end,
                    "faculty": email_key,
                    "room": room,
                    "section": section,
                }
            )

    seen_emails: set[str] = set()
    sem_num = {"III": 3, "V": 5, "VII": 7}[sem_roman]
    for row in rows:
        for cell in row:
            if isinstance(cell, str) and "@mygyanvihar.com" in cell.lower():
                email = cell.strip().lower()
                if email not in seen_emails:
                    seen_emails.add(email)
                    students.append({"email": email, "semester": sem_num, "section": section})

    return {
        "semester_roman": sem_roman,
        "semester_num": sem_num,
        "section": section,
        "room": room,
        "courses": list({c["code"]: c for c in courses.values()}.values()),
        "slots": slots,
        "students": students,
    }


def kind_is_lab(name: str) -> bool:
    u = name.upper()
    return "LAB" in u or "PRACTICAL" in u or "PRACT" in u


def parse_faculty_sheet(ws, faculty_key: str, default_room: str) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    time_cols: list[tuple[int, str, str]] = []
    header_idx = None
    for i, row in enumerate(rows):
        if row and isinstance(row[0], str) and "day" in row[0].lower():
            header_idx = i
            for col_idx, cell in enumerate(row):
                if col_idx == 0:
                    continue
                parsed = parse_time_header(str(cell) if cell else None)
                if parsed:
                    time_cols.append((col_idx, parsed[0], parsed[1]))
            break
    if header_idx is None:
        return []

    slots: list[dict] = []
    email_key = faculty_key
    for row in rows[header_idx + 1 :]:
        if not row or not row[0]:
            continue
        day_cell = str(row[0]).strip().lower()
        if day_cell not in DAYS:
            continue
        day = DAYS[day_cell]
        for col_idx, start, end in time_cols:
            if col_idx >= len(row):
                continue
            cell = row[col_idx]
            if not cell or not isinstance(cell, str):
                continue
            text = cell.strip()
            if is_skip_slot(text):
                continue
            name = clean_course_name(text)
            if not name:
                continue
            sem_roman = infer_sem_from_faculty_cell(name)
            slots.append(
                {
                    "raw_name": name,
                    "semester_roman": sem_roman,
                    "day": day,
                    "start": start,
                    "end": end,
                    "faculty": email_key,
                    "room": default_room,
                }
            )
    return slots


def infer_sem_from_faculty_cell(name: str) -> str:
    u = name.upper()
    if "7TH" in u or " VII" in u:
        return "VII"
    if "5TH" in u or " V SEM" in u or "BATCH A V" in u or "BATCH B" in u and "V SEM" in u:
        return "V"
    if "3RD" in u or " III" in u:
        return "III"
    if "ORTHO" in u and "BATCH A V" in u:
        return "V"
    if "EXERCISE THERAPY" in u:
        return "III"
    if "NEUROLOGICAL" in u or "CARDIOTORACIC" in u or "RESPIRATORY" in u:
        return "VII"
    return "V"


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Missing workbook: {XLSX}")

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    sem3 = parse_batch_sheet(wb["Sem 3rd batch a"], "III", "BPT-LT-1", "A")
    sem5 = parse_batch_sheet(wb["Sem 5th Batch A"], "V", "BPT-LT-2", "A")

    faculty_sheets = {
        "RG": parse_faculty_sheet(wb["RG"], "RG", "BPT-LT-1"),
        "PB": parse_faculty_sheet(wb["PB"], "PB", "BPT-LT-2"),
        "AS": parse_faculty_sheet(wb["AS"], "AS", "BPT-LT-3"),
    }

    # Merge course catalog across semesters
    all_courses: dict[str, dict] = {}
    code_by_name_sem: dict[tuple[str, str], str] = {}

    def register_course(course: dict, sem_roman: str) -> str:
        key = (course["name"].upper(), sem_roman)
        if key in code_by_name_sem:
            return code_by_name_sem[key]
        code = course["code"]
        code_by_name_sem[key] = code
        all_courses[code] = {**course, "semester": sem_roman}
        return code

    for block in (sem3, sem5):
        remapped_courses: dict[str, dict] = {}
        for c in block["courses"]:
            code = register_course(c, block["semester_roman"])
            remapped_courses[code] = all_courses[code]
            for slot in block["slots"]:
                if slot["course"] == c["code"]:
                    slot["course"] = code
        block["courses"] = list(remapped_courses.values())

    # Map faculty sheet slots to course codes
    faculty_slots: list[dict] = []
    code_set = set(all_courses)
    for fkey, fslots in faculty_sheets.items():
        for slot in fslots:
            name = slot["raw_name"]
            sem = slot["semester_roman"]
            key = (name.upper(), sem)
            if key not in code_by_name_sem:
                code = slug_code(name, sem, code_set)
                ctype = "LAB" if kind_is_lab(name) else "THEORY"
                all_courses[code] = {
                    "code": code,
                    "name": name,
                    "credits": 2 if ctype == "LAB" else 4,
                    "type": ctype,
                    "faculty": fkey,
                    "semester": sem,
                }
                code_by_name_sem[key] = code
            code = code_by_name_sem[key]
            faculty_slots.append(
                {
                    "course": code,
                    "semester": sem,
                    "day": slot["day"],
                    "start": slot["start"],
                    "end": slot["end"],
                    "faculty": fkey,
                    "room": slot["room"],
                }
            )

    students = sem3["students"] + sem5["students"]

    payload = {
        "academic_year": "2026-2027",
        "program_name": "BPT",
        "program_code": "BPT",
        "dept_name": "BPT",
        "faculty": FACULTY_EMAIL,
        "hod_email": "gaurav.agarwal@mygyanvihar.com",
        "mentorship_by_semester": {"3": "RG", "5": "RG"},
        "semesters": {
            "III": {
                "room": sem3["room"],
                "section": "A",
                "courses": sem3["courses"],
                "slots": sem3["slots"],
            },
            "V": {
                "room": sem5["room"],
                "section": "A",
                "courses": sem5["courses"],
                "slots": sem5["slots"],
            },
        },
        "faculty_timetable_slots": faculty_slots,
        "catalog": list(all_courses.values()),
        "students": students,
        "student_users_to_ensure": [
            {
                "email": s["email"],
                "semester": s["semester"],
                "section": s.get("section", "A"),
                "name": email_to_name(s["email"]),
            }
            for s in students
        ],
    }

    OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUT}")
    print(
        f"  courses={len(all_courses)} sem3_slots={len(sem3['slots'])} "
        f"sem5_slots={len(sem5['slots'])} faculty_slots={len(faculty_slots)} "
        f"students={len(students)}"
    )


def email_to_name(email: str) -> str:
    local = email.split("@")[0]
    parts = local.split(".")
    if len(parts) >= 2:
        return parts[0].title() + " " + parts[1].title()
    return local.title()


if __name__ == "__main__":
    main()
