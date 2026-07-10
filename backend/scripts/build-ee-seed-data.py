#!/usr/bin/env python3
"""Build ee-seed-data.json from EE Excel source files."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
OUT = Path(__file__).resolve().parent / "ee-seed-data.json"

DAYS = {
    "MONDAY": 1,
    "TUESDAY": 2,
    "WEDNESDAY": 3,
    "THURSDAY": 4,
    "FRIDAY": 5,
    "SATURDAY": 6,
}

SKIP = re.compile(
    r"weekly test|monday test|^\s*pt\s*$|library|vac|community|remedial|"
    r"employability|field project|industrial training|weekly activity|eqs|^\s*$",
    re.I,
)

COURSE_MAP = {
    "AE(HRT)": ("EE301T", "Applied Electronics", 4, "THEORY"),
    "AE Lab (HRT)": ("EE301P", "Applied Electronics Lab", 2, "LAB"),
    "EM (HRT)": ("EE302T", "Electrical Machines", 4, "THEORY"),
    "EM-1 (NKS)": ("EE303T", "Electrical Machines I", 4, "THEORY"),
    "EM-I LAB (NKS)": ("EE303P", "Electrical Machines I Lab", 2, "LAB"),
    "ECA (NF)": ("EE304T", "Electrical Circuit Analysis", 3, "THEORY"),
    "ECD LAB (NF)": ("EE304P", "Electrical Circuit Design Lab", 2, "LAB"),
    "ECO(RK)": ("EE305T", "Economics for Engineers", 2, "THEORY"),
    "PGP (RJ)": ("EE306T", "Professional Growth Program", 2, "THEORY"),
    "Microprocessor (NF)": ("EE501T", "Microprocessor", 4, "THEORY"),
    "Microprocessor Lab (NF)": ("EE501P", "Microprocessor Lab", 2, "LAB"),
    "Control System (NF)": ("EE502T", "Control System", 4, "THEORY"),
    "Control System Lab (NF)": ("EE502P", "Control System Lab", 2, "LAB"),
    "System Programming Lab (RJ)": ("EE503P", "System Programming Lab", 2, "LAB"),
    "MPE (RJ)": ("EE504T", "Microprocessor Applications", 3, "THEORY"),
    "PS-I (MKG)": ("EE505T", "Power System I", 4, "THEORY"),
    "PS-I LAB (MKG)": ("EE505P", "Power System I Lab", 2, "LAB"),
    "RES (HRT)": ("EE506T", "Renewable Energy Systems", 3, "THEORY"),
    "EE Class (NKS)": ("EE507T", "Electrical Engineering Class", 2, "THEORY"),
}


def norm_time(raw) -> tuple[str, str] | None:
    if raw is None:
        return None
    t = str(raw).strip().lower().replace(" ", "")
    m = re.match(r"(\d{1,2}):(\d{2})-(\d{1,2}):(\d{2})", t)
    if not m:
        return None
    sh, sm, eh, em = map(int, m.groups())

    def fix_hour(h: int) -> int:
        if 1 <= h <= 7:
            return h + 12
        return h

    sh, eh = fix_hour(sh), fix_hour(eh)
    return f"{sh:02d}:{sm:02d}", f"{eh:02d}:{em:02d}"


def faculty_key(cell: str) -> str:
    if "(RJ)" in cell or " (RJ)" in cell:
        return "ritu"
    if "(RK)" in cell or "(RK)" in cell:
        return "raj"
    if "HRT" in cell:
        return "raj"
    if "(NF)" in cell or "(NF)" in cell:
        return "ritu"
    if "(NKS)" in cell or "(NKS)" in cell:
        return "raj"
    if "(MKG)" in cell or "(MKG)" in cell:
        return "raj"
    return "ritu"


def clean_name(cell) -> str:
    s = str(cell).strip().split("\n")[0].strip()
    s = re.sub(r"\s*\([A-Z]+\)\s*", "", s)
    s = re.sub(r"\s*EE-[IVX]+", "", s, flags=re.I)
    return s.strip()


def resolve_course(name: str) -> tuple[str, str, int, str]:
    for key, meta in COURSE_MAP.items():
        if key.lower() in name.lower() or name.lower() in key.lower():
            return meta
    is_lab = "lab" in name.lower()
    code = "EE" + re.sub(r"[^A-Z0-9]", "", name.upper())[:6]
    code = (code + ("P" if is_lab else "T"))[:8]
    return code, name, 2 if is_lab else 3, "LAB" if is_lab else "THEORY"


def parse_student_sheet(path: Path, sheet: str, default_room: str) -> dict:
    ws = load_workbook(path, data_only=True)[sheet]
    room = default_room
    for r in range(1, 15):
        val = ws.cell(r, 1).value
        if not val:
            continue
        text = str(val)
        if "LT-" in text:
            m = re.search(r"LT-\d+", text)
            if m:
                room = m.group(0)
        if "MACHINE Lab" in text:
            room = "MACHINE Lab EC Block"

    times: dict[int, tuple[str, str]] = {}
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip().upper() == "TIME/DAY":
            for c in range(2, 10):
                nt = norm_time(ws.cell(r, c).value)
                if nt:
                    times[c] = nt
            break

    courses: dict[str, dict] = {}
    slots: list[dict] = []
    for r in range(1, ws.max_row + 1):
        day = str(ws.cell(r, 1).value or "").strip().upper()
        if day not in DAYS:
            continue
        for col, (start, end) in times.items():
            raw = ws.cell(r, col).value
            if raw is None:
                continue
            name = clean_name(raw)
            if not name or SKIP.search(name):
                continue
            fk = faculty_key(str(raw))
            code, cname, cred, ctype = resolve_course(name)
            courses[code] = {
                "code": code,
                "name": cname,
                "credits": cred,
                "type": ctype,
                "faculty": fk,
            }
            slots.append(
                {
                    "course": code,
                    "day": DAYS[day],
                    "start": start,
                    "end": end,
                    "faculty": fk,
                }
            )

    return {"room": room, "courses": list(courses.values()), "slots": slots}


def parse_ids(path: Path) -> list[dict]:
    ws = load_workbook(path, data_only=True)["B.Tech EE"]
    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        designation = ws.cell(r, 3).value
        email = ws.cell(r, 4).value
        if not email:
            continue
        email = str(email).strip().lower()
        des = str(designation or "").strip().lower()
        sem = None
        if "3" in des:
            sem = 3
        elif "5" in des:
            sem = 5
        if "hod" in des:
            role = "HOD"
        elif "faculty" in des:
            role = "Faculty"
        else:
            role = "Student"
        rows.append({"name": str(name).strip(), "email": email, "role": role, "semester": sem})
    return rows


def main() -> None:
    stt = ROOT / "EE (3 and 5 sem) TIME TABLE 2026.xlsx"
    ids = ROOT / "EE_Ids.xlsx"
    if not stt.exists() or not ids.exists():
        raise SystemExit("Missing EE source xlsx files at repo root")

    payload = {
        "academic_year": "2026-2027",
        "program_name": "B.Tech EE",
        "faculty": {
            "paresh": "paresh.jain@mygyanvihar.com",
            "ritu": "ritu.jain@mygyanvihar.com",
            "raj": "raj.kumar@mygyanvihar.com",
        },
        "credentials": parse_ids(ids),
        "semesters": {
            "III": parse_student_sheet(stt, "EE-III", "MACHINE Lab EC Block"),
            "V": parse_student_sheet(stt, "EE-V", "LT-18"),
        },
        "mentorship_by_semester": {"3": "raj", "5": "ritu"},
    }
    payload["students"] = [
        {"email": r["email"], "semester": r["semester"]}
        for r in payload["credentials"]
        if r["role"] == "Student" and r["semester"]
    ]

    OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
