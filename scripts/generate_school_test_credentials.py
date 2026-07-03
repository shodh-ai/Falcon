"""Generate SQL migration for school test credentials from Falcon Test IDs xlsx."""
from __future__ import annotations

import re
import sys
import uuid
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Falcon Test IDs (Schools).xlsx"
OUT = ROOT / "backend" / "migrations" / "20260703120000_school_test_credentials_seed.sql"

DOMAIN_FIXES = {
    "mygynavihar.com": "mygyanvihar.com",
    "mygyanviihar.com": "mygyanvihar.com",
    "mygyan.com": "mygyanvihar.com",
    "myyanvihar.com": "mygyanvihar.com",
}

SKIP_EMAILS = {
    "anushkagoyal079@gmail.com",
    "ramchandani.vanshika1204@gmail.com",
    "deepak2552709@mygyanvihar.com",
    "deepak255092@mygyanvihar.com",
}

SKIP_VALUES = {"NA", "N/A", "-", ""}

SHEET_DEPARTMENTS = {
    "CA": ("CA", "School of Commerce & Accountancy"),
    "ISBM": ("ISBM", "Institute of Business Management"),
    "Pharmacy": ("Pharmacy", "School of Pharmacy"),
    "Electrical Engg": ("Electrical Engg", "Department of Electrical Engineering"),
    "Clinical Psychology": ("Clinical Psychology", "Department of Clinical Psychology"),
    "SILS": ("SILS", "Suresh Gyan Vihar International Literacy School"),
    "C3WR": ("C3WR", "Centre for Climate Change and Water Research"),
    "Applied Sciences": ("Applied Sciences", "School of Applied Sciences"),
    "Mech Engg": ("Mech Engg", "Department of Mechanical Engineering"),
    "BPT": ("BPT", "Department of Physiotherapy"),
    "GCAD": ("GCAD", "Gyan Vihar Centre for Art and Design"),
    "Civil": ("Civil", "Department of Civil Engineering"),
    "Law": ("Law", "School of Law"),
    "Education": ("Education", "School of Education"),
    "Agriculture": ("Agriculture", "School of Agriculture"),
}

PWD_HASH = "$2b$10$3M.gdiob7z.LbjCitlN4DuM//mv4oNU1x1yGYD51wXFw30qVt8MoO"


def split_cell_lines(val):
    if val is None:
        return []
    return [ln.strip() for ln in re.split(r"[\r\n]+", str(val)) if ln.strip()]


def normalize_email(raw):
    if raw is None:
        return None
    text = str(raw).strip()
    if text.upper() in SKIP_VALUES:
        return None
    s = re.sub(r"\s+", "", text)
    match = re.search(r"([\w.\-+]+@([\w.\-]+))", s, re.I)
    if not match:
        return None
    local, domain = match.group(1).lower().split("@", 1)
    domain = DOMAIN_FIXES.get(domain, domain)
    return f"{local}@{domain}"


def clean_name(raw):
    name = re.sub(r"^\d+\.?\s*", "", str(raw or "?").strip())
    name = re.sub(r"\s+", " ", name).strip()
    return name or "Unknown"


def infer_role(designation: str) -> str:
    d = (designation or "").lower()
    if "dean" in d or "principal" in d:
        return "Dean"
    if "hod" in d or "head" in d:
        return "HOD"
    if "faculty" in d or "professor" in d or "member" in d:
        return "Faculty"
    return "Student"


def parse_semester(designation: str) -> int | None:
    if not designation:
        return None
    match = re.search(r"(\d+)\s*(?:st|nd|rd|th)?\s*sem", designation, re.I)
    return int(match.group(1)) if match else None


def enrollment_from_email(email: str) -> str:
    local = email.split("@", 1)[0]
    match = re.search(r"(\d{5,})", local)
    if match:
        return match.group(1)
    return local.replace(".", "").upper()[:20]


def sql_str(value: str) -> str:
    return value.replace("'", "''")


def extract_sheet(ws):
    rows = []
    for _, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = list(row)
        if not any(v is not None and str(v).strip() for v in vals):
            continue
        rows.append(vals)

    start = 0
    for idx, vals in enumerate(rows):
        joined = " ".join(str(v) for v in vals if v).lower()
        if "email" in joined or "gyan vihar" in joined:
            start = idx + 1
            break

    people = []
    for vals in rows[start:]:
        joined = " ".join(str(v) for v in vals if v).lower()
        if joined.startswith("s. no") or joined.startswith("s.no"):
            continue
        while len(vals) < 4:
            vals.append(None)
        _, name, desig, email_col = vals[0], vals[1], vals[2], vals[3]
        name_lines = split_cell_lines(name)
        email_lines = split_cell_lines(email_col)
        if len(name_lines) > 1 or len(email_lines) > 1:
            count = max(len(name_lines), len(email_lines), 1)
            names = name_lines or ["?"] * count
            emails_raw = email_lines or [""] * count
            while len(names) < count:
                names.append("?")
            while len(emails_raw) < count:
                emails_raw.append("")
            pairs = zip(names, emails_raw)
        else:
            pairs = [(name, email_col)]

        for nm, raw_email in pairs:
            email = normalize_email(raw_email)
            if not email or email in SKIP_EMAILS:
                continue
            people.append(
                {
                    "name": clean_name(nm),
                    "designation": str(desig or "").strip(),
                    "email": email,
                    "role": infer_role(str(desig or "")),
                    "semester": parse_semester(str(desig or "")),
                }
            )
    return people


def stable_user_id(email: str) -> str:
    return str(uuid.uuid5(uuid.NAMESPACE_DNS, f"sgvu-school-test:{email}"))


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else XLSX
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

    seen_emails: set[str] = set()
    users: list[dict] = []
    dept_people: dict[str, list[str]] = {}

    for sheet in wb.sheetnames:
        if sheet not in SHEET_DEPARTMENTS:
            continue
        dept_name, dept_desc = SHEET_DEPARTMENTS[sheet]
        for person in extract_sheet(wb[sheet]):
            email = person["email"]
            if email in seen_emails:
                continue
            seen_emails.add(email)
            person["sheet"] = sheet
            person["dept_name"] = dept_name
            person["dept_desc"] = dept_desc
            person["user_id"] = stable_user_id(email)
            users.append(person)
            dept_people.setdefault(dept_name, []).append(email)

    faculty = [u for u in users if u["role"] != "Student"]
    students = [u for u in users if u["role"] == "Student"]

    lines: list[str] = []
    lines.append("-- School test credentials from Falcon Test IDs (Schools).xlsx")
    lines.append("-- Default password: password123")
    lines.append("-- Excludes 4 special cases (2 gmail Clinical Psychology, 2 Agriculture deepak*)")
    lines.append("-- Unique persons only; cross-sheet duplicates use first sheet department")
    lines.append(f"-- Total users: {len(users)} ({len(faculty)} staff, {len(students)} students)")
    lines.append("")
    lines.append("DO $$")
    lines.append("BEGIN")
    for role in ("Student", "Faculty", "HOD", "Dean"):
        lines.append(f"  IF NOT EXISTS (SELECT 1 FROM roles WHERE role_name = '{role}') THEN")
        lines.append(f"    INSERT INTO roles (role_name, description)")
        lines.append(f"    VALUES ('{role}', 'Application role for {role} portal access');")
        lines.append("  END IF;")
    lines.append("END $$;")
    lines.append("")

    for dept_name, (_, dept_desc) in SHEET_DEPARTMENTS.items():
        lines.append(
            f"INSERT INTO departments (dept_name, description)\n"
            f"VALUES ('{sql_str(dept_name)}', '{sql_str(dept_desc)}')\n"
            f"ON CONFLICT (dept_name) DO NOTHING;"
        )
    lines.append("")

    def values_rows(records, include_enrollment=False):
        rows = []
        for u in records:
            if include_enrollment:
                enr = enrollment_from_email(u["email"])
                sem = u["semester"] if u["semester"] is not None else "NULL"
                rows.append(
                    f"    ('{u['user_id']}'::uuid, '{sql_str(u['name'])}', "
                    f"'{sql_str(u['email'])}', '{sql_str(u['dept_name'])}', "
                    f"'{sql_str(u['role'])}', '{sql_str(enr)}', {sem})"
                )
            else:
                rows.append(
                    f"    ('{u['user_id']}'::uuid, '{sql_str(u['name'])}', "
                    f"'{sql_str(u['email'])}', '{sql_str(u['dept_name'])}', "
                    f"'{sql_str(u['role'])}')"
                )
        return ",\n".join(rows)

    if faculty:
        lines.append("-- ---------------------------------------------------------------------------")
        lines.append("-- STAFF (Faculty / HOD / Dean)")
        lines.append("-- ---------------------------------------------------------------------------")
        lines.append("WITH tenant AS (")
        lines.append("  SELECT tenant_id FROM public.tenants WHERE subdomain = 'sgvu' LIMIT 1")
        lines.append("),")
        lines.append("pwd AS (")
        lines.append(f"  SELECT '{PWD_HASH}'::varchar AS hash")
        lines.append("),")
        lines.append("seed_staff AS (")
        lines.append("  SELECT * FROM (VALUES")
        lines.append(values_rows(faculty))
        lines.append("  ) AS s(user_id, name, email, dept_name, role_name)")
        lines.append(")")
        lines.append("INSERT INTO users (")
        lines.append("  user_id, tenant_id, name, official_email, role_id, dept_id,")
        lines.append("  password_hash, is_active, onboarding_status, onboarding_profile")
        lines.append(")")
        lines.append("SELECT")
        lines.append("  s.user_id, t.tenant_id, s.name, s.email, r.role_id, d.dept_id,")
        lines.append("  p.hash, true, 'PENDING_PASSWORD_RESET', '{}'::jsonb")
        lines.append("FROM seed_staff s")
        lines.append("CROSS JOIN tenant t")
        lines.append("CROSS JOIN pwd p")
        lines.append("JOIN departments d ON d.dept_name = s.dept_name")
        lines.append("JOIN roles r ON r.role_name = s.role_name")
        lines.append("ON CONFLICT (tenant_id, official_email) DO UPDATE SET")
        lines.append("  name = EXCLUDED.name,")
        lines.append("  role_id = EXCLUDED.role_id,")
        lines.append("  dept_id = EXCLUDED.dept_id,")
        lines.append("  password_hash = EXCLUDED.password_hash,")
        lines.append("  is_active = true,")
        lines.append("  onboarding_status = 'PENDING_PASSWORD_RESET',")
        lines.append("  onboarding_profile = '{}'::jsonb;")
        lines.append("")
        lines.append("INSERT INTO user_roles (user_id, role_id, is_primary)")
        lines.append("SELECT u.user_id, u.role_id, true")
        lines.append("FROM users u")
        lines.append("WHERE u.user_id IN (")
        lines.append(",\n".join(f"  '{u['user_id']}'::uuid" for u in faculty))
        lines.append(")")
        lines.append("ON CONFLICT (user_id, role_id) DO UPDATE SET is_primary = EXCLUDED.is_primary;")
        lines.append("")

    if students:
        lines.append("-- ---------------------------------------------------------------------------")
        lines.append("-- STUDENTS")
        lines.append("-- ---------------------------------------------------------------------------")
        lines.append("WITH tenant AS (")
        lines.append("  SELECT tenant_id FROM public.tenants WHERE subdomain = 'sgvu' LIMIT 1")
        lines.append("),")
        lines.append("pwd AS (")
        lines.append(f"  SELECT '{PWD_HASH}'::varchar AS hash")
        lines.append("),")
        lines.append("seed_students AS (")
        lines.append("  SELECT * FROM (VALUES")
        lines.append(values_rows(students, include_enrollment=True))
        lines.append("  ) AS s(user_id, name, email, dept_name, role_name, enrollment_no, semester_num)")
        lines.append(")")
        lines.append("INSERT INTO users (")
        lines.append("  user_id, tenant_id, name, official_email, role_id, dept_id,")
        lines.append("  password_hash, is_active, onboarding_status, onboarding_profile")
        lines.append(")")
        lines.append("SELECT")
        lines.append("  s.user_id, t.tenant_id, s.name, s.email, r.role_id, d.dept_id,")
        lines.append("  p.hash, true, 'PENDING_PASSWORD_RESET', '{}'::jsonb")
        lines.append("FROM seed_students s")
        lines.append("CROSS JOIN tenant t")
        lines.append("CROSS JOIN pwd p")
        lines.append("JOIN departments d ON d.dept_name = s.dept_name")
        lines.append("JOIN roles r ON r.role_name = 'Student'")
        lines.append("ON CONFLICT (tenant_id, official_email) DO UPDATE SET")
        lines.append("  name = EXCLUDED.name,")
        lines.append("  role_id = EXCLUDED.role_id,")
        lines.append("  dept_id = EXCLUDED.dept_id,")
        lines.append("  password_hash = EXCLUDED.password_hash,")
        lines.append("  is_active = true,")
        lines.append("  onboarding_status = 'PENDING_PASSWORD_RESET',")
        lines.append("  onboarding_profile = '{}'::jsonb;")
        lines.append("")
        lines.append("INSERT INTO user_roles (user_id, role_id, is_primary)")
        lines.append("SELECT u.user_id, u.role_id, true")
        lines.append("FROM users u")
        lines.append("WHERE u.user_id IN (")
        lines.append(",\n".join(f"  '{u['user_id']}'::uuid" for u in students))
        lines.append(")")
        lines.append("ON CONFLICT (user_id, role_id) DO UPDATE SET is_primary = EXCLUDED.is_primary;")
        lines.append("")
        lines.append("WITH seed_students AS (")
        lines.append("  SELECT * FROM (VALUES")
        lines.append(values_rows(students, include_enrollment=True))
        lines.append("  ) AS s(user_id, name, email, dept_name, role_name, enrollment_no, semester_num)")
        lines.append(")")
        lines.append("INSERT INTO student_profiles (")
        lines.append("  tenant_id, user_id, enrollment_no, enrollment_number, admission_number,")
        lines.append("  batch, nationality, admission_status, status")
        lines.append(")")
        lines.append("SELECT")
        lines.append("  u.tenant_id, u.user_id, s.enrollment_no, s.enrollment_no, s.enrollment_no,")
        lines.append("  d.dept_name, 'Indian', 'ACTIVE', 'ACTIVE'")
        lines.append("FROM seed_students s")
        lines.append("JOIN users u ON u.user_id = s.user_id")
        lines.append("JOIN departments d ON d.dept_name = s.dept_name")
        lines.append("ON CONFLICT (user_id) DO UPDATE SET")
        lines.append("  tenant_id = EXCLUDED.tenant_id,")
        lines.append("  enrollment_no = EXCLUDED.enrollment_no,")
        lines.append("  enrollment_number = EXCLUDED.enrollment_number,")
        lines.append("  admission_number = EXCLUDED.admission_number,")
        lines.append("  batch = EXCLUDED.batch,")
        lines.append("  updated_at = NOW();")

    OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(f"Generated {OUT}")
    print(f"Users: {len(users)} ({len(faculty)} staff, {len(students)} students)")
    print(f"Excluded special cases: {len(SKIP_EMAILS)}")
    print("\nBy department:")
    for dept_name in SHEET_DEPARTMENTS:
        count = len(dept_people.get(dept_name, []))
        if count:
            print(f"  {dept_name:25} {count:3}")


if __name__ == "__main__":
    main()
