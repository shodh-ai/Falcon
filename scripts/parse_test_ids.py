"""Parse Falcon Test IDs xlsx with multiline cells, space strip, typo fixes."""
import re
import sys
from collections import Counter

import openpyxl

DOMAIN_FIXES = {
    "mygynavihar.com": "mygyanvihar.com",
    "mygyanviihar.com": "mygyanvihar.com",
    "mygyan.com": "mygyanvihar.com",
    "myyanvihar.com": "mygyanvihar.com",
}

SKIP_VALUES = {"NA", "N/A", "-", ""}


def split_cell_lines(val):
    if val is None:
        return []
    return [ln.strip() for ln in re.split(r"[\r\n]+", str(val)) if ln.strip()]


def normalize_email(raw):
    if raw is None:
        return None, "empty"
    text = str(raw).strip()
    if text.upper() in SKIP_VALUES:
        return None, "skip_na"

    s = re.sub(r"\s+", "", text)
    match = re.search(r"([\w.\-+]+@([\w.\-]+))", s, re.I)
    if not match:
        return None, "no_email"

    local, domain = match.group(1).lower().split("@", 1)
    orig_domain = domain
    domain = DOMAIN_FIXES.get(domain, domain)
    fixed = f"{local}@{domain}"
    note = "typo_fixed" if orig_domain != domain else "ok"
    return fixed, note


def extract_from_sheet(ws):
    rows_data = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = list(row)
        if not any(v is not None and str(v).strip() for v in vals):
            continue
        rows_data.append((i, vals))

    people = []
    skipped = []
    typo_fixes = []

    start = 0
    for idx, (_, vals) in enumerate(rows_data):
        joined = " ".join(str(v) for v in vals if v).lower()
        if "email" in joined or "gyan vihar" in joined:
            start = idx + 1
            break

    for row_idx, vals in rows_data[start:]:
        joined = " ".join(str(v) for v in vals if v).lower()
        if joined.startswith("s. no") or joined.startswith("s.no"):
            continue

        while len(vals) < 4:
            vals.append(None)
        sno, name, desig, email_col = vals[0], vals[1], vals[2], vals[3]

        name_lines = split_cell_lines(name)
        email_lines = split_cell_lines(email_col)

        if len(name_lines) > 1 or len(email_lines) > 1:
            count = max(len(name_lines), len(email_lines), 1)
            names = name_lines if name_lines else ["?"] * count
            emails_raw = email_lines if email_lines else [""] * count
            while len(names) < count:
                names.append("?")
            while len(emails_raw) < count:
                emails_raw.append("")
            for nm, raw_email in zip(names, emails_raw):
                email, status = normalize_email(raw_email)
                if email:
                    if status == "typo_fixed":
                        typo_fixes.append((email, raw_email))
                    people.append(
                        {
                            "row": row_idx,
                            "name": nm,
                            "desig": desig,
                            "email": email,
                            "status": status,
                        }
                    )
                elif raw_email:
                    skipped.append(
                        {"row": row_idx, "name": nm, "raw": raw_email, "reason": status}
                    )
        else:
            email, status = normalize_email(email_col)
            if email:
                if status == "typo_fixed":
                    typo_fixes.append((email, str(email_col)))
                people.append(
                    {
                        "row": row_idx,
                        "name": name_lines[0] if name_lines else name,
                        "desig": desig,
                        "email": email,
                        "status": status,
                    }
                )
            elif email_col and str(email_col).strip().upper() not in SKIP_VALUES:
                skipped.append(
                    {"row": row_idx, "name": name, "raw": email_col, "reason": status}
                )

    by_email = {}
    dupes = []
    for person in people:
        if person["email"] in by_email:
            dupes.append((person["email"], by_email[person["email"]], person))
        else:
            by_email[person["email"]] = person

    return list(by_email.values()), skipped, typo_fixes, dupes


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else "Falcon Test IDs (Schools).xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    print("=" * 70)
    print("FALCON TEST IDs - FULL RE-READ")
    print("Rules: split multiline cells | strip spaces | fix typo domains")
    print("Domain fixes: mygynavihar, mygyanviihar, mygyan.com -> mygyanvihar.com")
    print("=" * 70)

    sheet_counts = {}
    grand_total = 0
    all_emails = []
    all_skipped = []
    all_typos = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        unique, skipped, typos, dupes = extract_from_sheet(ws)
        sheet_counts[sheet] = len(unique)
        grand_total += len(unique)
        all_emails.extend((sheet, p["email"]) for p in unique)

        print(f"\n## {sheet}: {len(unique)} emails")
        if typos:
            print("  Typo fixes:")
            for fixed, raw in typos:
                print(f"    {raw} -> {fixed}")
        if dupes:
            print(f"  Duplicate emails in sheet: {len(dupes)}")
            for email, _, _ in dupes:
                print(f"    {email}")
        if skipped:
            print(f"  Skipped ({len(skipped)}):")
            for item in skipped:
                print(f"    row {item['row']}: {item['raw']} ({item['reason']})")
        for idx, person in enumerate(unique, 1):
            flag = " [fixed]" if person["status"] == "typo_fixed" else ""
            print(f"  {idx:2}. {person['email']}{flag}")

        all_skipped.extend((sheet, item) for item in skipped)
        all_typos.extend((sheet, fixed, raw) for fixed, raw in typos)

    print("\n" + "=" * 70)
    print("SUMMARY BY SHEET")
    print("=" * 70)
    for sheet, count in sheet_counts.items():
        print(f"  {sheet:25} {count:3} emails")
    print("-" * 70)
    print(f"  {'TOTAL':25} {grand_total:3} emails")

    email_counter = Counter(email for _, email in all_emails)
    cross = [(email, n) for email, n in email_counter.items() if n > 1]
    unique_global = len(email_counter)

    print(f"\nGrand unique emails (across all sheets): {unique_global}")
    if cross:
        print(f"\nCross-sheet duplicate emails ({len(cross)}):")
        sheet_map = {}
        for sheet in wb.sheetnames:
            unique, _, _, _ = extract_from_sheet(wb[sheet])
            for person in unique:
                sheet_map.setdefault(person["email"], []).append(sheet)
        for email, n in sorted(cross, key=lambda x: -x[1]):
            print(f"  {email} -> {', '.join(sheet_map[email])}")

    if all_skipped:
        print(f"\nAll skipped entries ({len(all_skipped)}):")
        for sheet, item in all_skipped:
            print(f"  [{sheet}] row {item['row']}: {item['raw']} ({item['reason']})")

    if all_typos:
        print(f"\nAll typo fixes ({len(all_typos)}):")
        for sheet, fixed, raw in all_typos:
            print(f"  [{sheet}] {raw} -> {fixed}")


if __name__ == "__main__":
    main()
